import openpyxl

file = r"C:\Users\deepa\Downloads\FSI-2023-DOWNLOAD.xlsx"
workbook = openpyxl.load_workbook(file)

sheet = workbook["Sheet1"]

rows = sheet.max_row
column = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value)

    print()

