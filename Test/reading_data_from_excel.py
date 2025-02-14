import openpyxl

file = r"C:\Users\deepa\OneDrive\Desktop\Production Version Combined Zipcodes-3.xlsx"

workbook = openpyxl.load_workbook(file)

sheet = workbook["FWA-FIOS customers"]
rows  = sheet.max_row
cols = sheet.max_column


lst  =[]

for r in range(1,rows+1):
    lst.append(sheet.cell(r,2).value)

print(lst)

